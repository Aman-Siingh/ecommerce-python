import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def edit_file(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_value = cell.value * 0.9
        corrected_value_cell = sheet.cell(row, 5)
        corrected_value_cell.value = corrected_value

    Ref = Reference(sheet, min_col=5, min_row=2, max_row=sheet.max_row, max_col=5)
    Value = BarChart()
    Value.add_data(Ref)
    sheet.add_chart(Value, "f2")

    wb.save(filename)
